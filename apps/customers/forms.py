from django import forms
from django.utils import timezone

from apps.customers.models import Customer
from apps.sims.models import SIM
from apps.tariffs.models import Tariff
import csv
import io
from openpyxl import load_workbook
import uuid


class CustomerForm(forms.ModelForm):
    """
    Расширенная форма для работы с абонентами в ASMAN CRM.
    Позволяет одновременно создать карточку и при необходимости подключить номер.
    """

    sim_card = forms.ModelChoiceField(
        queryset=SIM.objects.none(),
        required=False,
        label='Выбрать номер (MSISDN)',
        help_text='Номер из пула заранее выпущенных SIM-карт',
    )
    tariff = forms.ModelChoiceField(
        queryset=Tariff.objects.none(),
        required=False,
        label='Тарифный план',
        help_text='Необходимо выбрать тариф для подключения номера',
    )

    def __init__(self, *args, **kwargs):
        self.enable_sim_assignment = kwargs.pop('enable_sim_assignment', True)
        super().__init__(*args, **kwargs)
        self.created_contracts = []

        base_input = (
            "mt-2 block w-full rounded-xl border border-gray-200 bg-white/80 px-3 py-2 "
            "text-sm text-gray-900 placeholder:text-gray-400 shadow-sm focus:border-primary-500 "
            "focus:ring-4 focus:ring-primary-100 transition"
        )

        for field_name in self.fields:
            field = self.fields[field_name]
            widget = field.widget
            if isinstance(widget, (forms.TextInput, forms.EmailInput, forms.NumberInput)):
                widget.attrs.setdefault('class', base_input)
            elif isinstance(widget, forms.Select):
                widget.attrs.setdefault('class', base_input + " pr-10")
            elif isinstance(widget, forms.Textarea):
                widget.attrs.setdefault('class', base_input)

        self.fields['customer_type'].label = 'Тип клиента'
        self.fields['customer_type'].widget.attrs.setdefault('x-model', 'type')
        self.fields['organization_name'].label = 'Название организации'
        self.fields['organization_code'].label = 'ИНН/БИН'
        self.fields['organization_name'].required = False
        self.fields['organization_code'].required = False

        self.fields['document_type'].required = False
        self.fields['passport_series'].required = False
        self.fields['passport_number'].required = False

        self.fields['first_name'].label = 'Имя'
        self.fields['first_name'].required = False
        self.fields['last_name'].label = 'Фамилия'
        self.fields['last_name'].required = False
        self.fields['patronymic'].label = 'Отчество'
        self.fields['patronymic'].required = False
        self.fields['birth_date'].label = 'Дата рождения'
        self.fields['birth_date'].required = False
        self.fields['birth_date'].widget = forms.DateInput(
            attrs={'type': 'date', 'class': base_input}
        )
        self.fields['phone'].help_text = 'Формат: +996 XXX XXX XXX'
        self.fields['phone'].widget.attrs.setdefault('placeholder', '+996 700 000 000')
        self.fields['email'].required = False
        self.fields['address'].widget = forms.Textarea(
            attrs={'rows': 3, 'class': base_input}
        )
        self.fields['address'].label = 'Адрес проживания'
        self.fields['inn'].label = 'ИНН (при наличии)'
        self.fields['inn'].required = False
        self.fields['document_type'].required = False
        self.fields['passport_series'].required = False
        self.fields['passport_number'].required = False

        if self.enable_sim_assignment:
            self.fields['sim_card'].queryset = SIM.objects.filter(status='free').order_by('msisdn')
            self.fields['tariff'].queryset = Tariff.objects.filter(is_active=True).order_by('name')
            self.fields['sim_card'].widget.attrs.update({
                'class': base_input,
                'data-controller': 'sim-picker',
            })
            self.fields['tariff'].widget.attrs.update({'class': base_input})
            self.fields['phone'].required = False
            self.fields['phone'].widget.attrs['placeholder'] = 'Будет подставлен из выбранной SIM'
        else:
            self.fields.pop('sim_card')
            self.fields.pop('tariff')
            self.fields['phone'].required = True

    class Meta:
        model = Customer
        fields = [
            'customer_type',
            'organization_name',
            'organization_code',
            'first_name',
            'last_name',
            'patronymic',
            'birth_date',
            'document_type',
            'passport_series',
            'passport_number',
            'inn',
            'phone',
            'email',
            'address',
            'status',
        ]
        labels = {
            'customer_type': 'Тип клиента',
            'organization_name': 'Название организации',
            'organization_code': 'ИНН/БИН',
            'passport_series': 'Серия документа',
            'passport_number': 'Номер документа',
            'status': 'Статус клиента',
        }

    def clean(self):
        cleaned = super().clean()
        customer_type = cleaned.get('customer_type', 'individual')

        if customer_type == 'organization':
            if not cleaned.get('organization_name'):
                self.add_error('organization_name', 'Введите название организации')
            # Для организаций можно не заполнять ФИО
            if not cleaned.get('first_name'):
                cleaned['first_name'] = 'Org'
            if not cleaned.get('last_name'):
                cleaned['last_name'] = cleaned.get('organization_name') or 'Org'
            if not cleaned.get('document_type'):
                cleaned['document_type'] = self.instance.document_type or 'passport'
            if not cleaned.get('passport_series'):
                org_code = (cleaned.get('organization_code') or 'ORG')[:10]
                cleaned['passport_series'] = org_code if org_code else 'ORG'
            if not cleaned.get('passport_number'):
                cleaned['passport_number'] = uuid.uuid4().hex[:8].upper()
            if not cleaned.get('phone'):
                generated = self._generate_unique_phone()
                cleaned['phone'] = generated
            self.instance.passport_series = cleaned['passport_series']
            self.instance.passport_number = cleaned['passport_number']
            self.instance.document_type = cleaned['document_type']
            self.instance.phone = cleaned['phone']
        else:
            if not cleaned.get('first_name'):
                self.add_error('first_name', 'Введите имя')
            if not cleaned.get('last_name'):
                self.add_error('last_name', 'Введите фамилию')
            for field, label in [('document_type', 'тип документа'), ('passport_series', 'серия документа'), ('passport_number', 'номер документа')]:
                if not cleaned.get(field):
                    self.add_error(field, f'Укажите {label}')

        if not self.enable_sim_assignment:
            if customer_type != 'organization' and not cleaned.get('phone'):
                self.add_error('phone', 'Телефон обязателен для физ. лиц')
            return cleaned

        sim = cleaned.get('sim_card')
        tariff = cleaned.get('tariff')
        if sim and sim.status != 'free':
            self.add_error('sim_card', 'Номер уже назначен другому договору.')
        if sim and not tariff:
            self.add_error('tariff', 'Выберите тариф, чтобы подключить выбранный номер.')
        if sim and not cleaned.get('phone'):
            cleaned['phone'] = sim.msisdn
            self.instance.phone = sim.msisdn
        return cleaned

    def _generate_unique_phone(self):
        from random import randint
        phone = f"+996{randint(0, 999999999):09d}"
        while Customer.objects.filter(phone=phone).exists():
            phone = f"+996{randint(0, 999999999):09d}"
        return phone


class CustomerSimAssignForm(forms.Form):
    sims = forms.ModelMultipleChoiceField(
        queryset=SIM.objects.none(),
        label='Свободные SIM',
        help_text='Выберите одну или несколько SIM-карт для подключения'
    )
    tariff = forms.ModelChoiceField(
        queryset=Tariff.objects.filter(is_active=True).order_by('name'),
        label='Тариф',
        help_text='Будет применён для всех выбранных SIM'
    )

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        base = (
            "mt-2 block w-full rounded-xl border border-gray-200 bg-white/80 px-3 py-2 text-sm "
            "text-gray-900 placeholder:text-gray-400 shadow-sm focus:border-primary-500 "
            "focus:ring-4 focus:ring-primary-100 transition"
        )
        self.fields['sims'].queryset = SIM.objects.filter(status='free').order_by('msisdn')
        self.fields['sims'].widget.attrs.update({'class': base + ' min-h-[200px]'})
        self.fields['sims'].widget.attrs['multiple'] = 'multiple'
        self.fields['tariff'].widget.attrs.update({'class': base})


class OrganizationBulkSimUploadForm(forms.Form):
    organization = forms.ModelChoiceField(
        queryset=Customer.objects.filter(customer_type='organization').order_by('organization_name'),
        label='Организация'
    )
    file = forms.FileField(
        label='Файл с SIM (CSV/XLSX)',
        help_text='Используйте шаблон, который можно скачать ниже'
    )

    def clean_file(self):
        uploaded = self.cleaned_data['file']
        name = uploaded.name.lower()
        if not (name.endswith('.csv') or name.endswith('.xlsx')):
            raise forms.ValidationError('Поддерживаются только CSV или XLSX')
        return uploaded

    def parse_rows(self):
        uploaded = self.cleaned_data['file']
        name = uploaded.name.lower()
        uploaded.seek(0)
        if name.endswith('.csv'):
            text = uploaded.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(text))
            return list(reader)
        wb = load_workbook(uploaded, read_only=True)
        ws = wb.active
        headers = [str(cell.value).strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for raw in ws.iter_rows(min_row=2, values_only=True):
            row = {}
            for header, value in zip(headers, raw):
                row[header] = value
            rows.append(row)
        return rows

    def save(self, commit=True):
        instance = super().save(commit=commit)
        self.assign_contract(instance)
        return instance

    def assign_contract(self, customer):
        """
        Создаёт договор и активирует SIM, если пользователь выбрал номер и тариф.
        """
        if not self.enable_sim_assignment:
            return

        sim = self.cleaned_data.get('sim_card')
        tariff = self.cleaned_data.get('tariff')
        if not (sim and tariff):
            return

        from apps.contracts.models import Contract

        contract = Contract.objects.create(
            customer=customer,
            tariff=tariff,
            signed_date=timezone.now().date(),
            status='draft',
            notes='Автоматическое подключение при регистрации абонента в ASMAN CRM.',
        )
        contract.activate(sim)
        self.created_contracts.append(contract)
